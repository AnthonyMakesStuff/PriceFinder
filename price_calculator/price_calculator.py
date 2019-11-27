"""
Price Calculator
Created by Anthony Provenza
Copyright 2019 Anthony Provenza
"""

import os
import string
import sys

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook, load_workbook


class Constant:
    """
    These are the constant value in the code, that are not changed
    """
    working_products_decimal = 0.8
    estimated_quantity_demand = {"NA": 0,
                                 "0-10": 0.35,
                                 "10-25": 0.4,
                                 "25-50": 0.5,
                                 "50-100": 0.6,
                                 "100-250": 0.7,
                                 "250+": 0.8}


class Currency:
    # noinspection SpellCheckingInspection
    def __init__(self,
                 currency_name: str,
                 currency_name_abrv: str,
                 currency_symbol: str):
        """
        Create and store a currency

        :param currency_name: The name of the currency (eg: Great British Pounds)
        :param currency_name_abrv: The abreviated letters (eg: GBP)
        :param currency_symbol: The symbol of the currency (eg: £)
        """
        self.currency_name = currency_name
        self.currency_name_abrv = currency_name_abrv
        self.currency_symbol = currency_symbol

    def __str__(self):
        return "{} ({})".format(self.currency_name, self.currency_symbol)


class ProductData:
    # noinspection PyPep8Naming
    def __init__(self,
                 ASIN: str,
                 ASIN_location: str,
                 quantity: int,
                 condition: str,
                 description: str,
                 website_amazon_price: float,
                 price_currency: Currency,
                 current_amazon_price: float) -> None:
        """
        Creates an object to store the data of a product

        :param ASIN: The Amazon Standard Identification Number
        :param ASIN_location: The location that this ASIN applies to, as the ASINs very
        :param quantity: The quantity of the specified product
        :param condition: The condition of the product
        :param description: The name of the product / the description of it
        :param website_amazon_price: The price that the website claims it to be
        :param price_currency: The currency the price is in
        :param current_amazon_price: The actual current Amazon price
        :return: None
        """
        self.ASIN = ASIN
        self.ASIN_location = ASIN_location
        self.quantity = quantity
        self.condition = condition
        self.description = description
        self.website_amazon_price = website_amazon_price
        self.price_currency = price_currency
        self.current_amazon_price = current_amazon_price

    def data_out(self) -> list:
        """
        Get the data of the class as a list

        :return: A list of the data of the product
        """
        return [self.ASIN,
                self.ASIN_location,
                self.quantity,
                self.condition,
                self.description,
                self.website_amazon_price,
                self.price_currency,
                self.current_amazon_price]

    def __repr__(self) -> str:
        return "ProductData(ASIN: \"{}\", ASIN_location: \"{}\", quantity: {}, condition: \"{}\", description: \"{}\"" \
               ", website_amazon_price: {}, price_currency: {}, current_amazon_price: {})".format(
                self.ASIN,
                self.ASIN_location,
                self.quantity,
                self.condition,
                self.description,
                self.website_amazon_price,
                self.price_currency,
                self.current_amazon_price)


class Analysis:
    """
    This class contains functions to analyse data
    """

    def __init__(self) -> None:
        pass

    @staticmethod
    def analyze_amazon_prices(data: list, column_index: int = 0) -> dict:
        analysed_data = {"NA": [0, 0],
                         "0-10": [0, 0],
                         "10-25": [0, 0],
                         "25-50": [0, 0],
                         "50-100": [0, 0],
                         "100-250": [0, 0],
                         "250+": [0, 0]}
        for product in data:
            if (not product.current_amazon_price):
                analysed_data["NA"][0] += 1
                analysed_data["NA"][1] += 0
            elif (product.current_amazon_price <= 10):
                analysed_data["0-10"][0] += 1
                analysed_data["0-10"][1] += product.current_amazon_price
            elif (10 < product.current_amazon_price <= 25):
                analysed_data["10-25"][0] += 1
                analysed_data["10-25"][1] += product.current_amazon_price
            elif (25 < product.current_amazon_price <= 50):
                analysed_data["25-50"][0] += 1
                analysed_data["25-50"][1] += product.current_amazon_price
            elif (50 < product.current_amazon_price <= 100):
                analysed_data["50-100"][0] += 1
                analysed_data["50-100"][1] += product.current_amazon_price
            elif (100 < product.current_amazon_price <= 250):
                analysed_data["100-250"][0] += 1
                analysed_data["100-250"][1] += product.current_amazon_price
            elif (250 < product.current_amazon_price):
                analysed_data["250+"][0] += 1
                analysed_data["250+"][1] += product.current_amazon_price
        return analysed_data


class WebsiteAnalysis:
    lot_price = 0
    xlsx_headings = ["ASIN",
                     "ASIN Location",
                     "Quantity",
                     "Condition",
                     "Description",
                     "Expected Amazon Price",
                     "Currency",
                     "Current Amazon Price"]
    product_data = list()
    price_analysis = dict()

    def __init__(self, url: str) -> None:
        """
        Analyse a website's data

        :param url: The URL of the website to analyse
        """
        self.url = url

    def analyse(self,
                data_location: str = None,
                website_currency: Currency = None) -> None:
        pass

    def to_xlsx(self,
                xlsx_location: str = "data.xlsx") -> None:
        workbook = Workbook()
        products_worksheet = workbook.active
        products_worksheet.title = "Products"
        analysis_worksheet = workbook.create_sheet(title="Analysis")

        # Products Worksheet
        for column_number, heading in enumerate(self.xlsx_headings):
            position = "{}{}".format(list(string.ascii_uppercase)[column_number],
                                     1)
            products_worksheet[position] = heading
        for row_number, row_data in enumerate(self.product_data):
            for column_number, column_data in enumerate(row_data.data_out()):
                position = "{}{}".format(list(string.ascii_uppercase)[column_number],
                                         row_number + 2)
                if (type(column_data) == Currency):
                    products_worksheet[position] = column_data.__str__()
                else:
                    products_worksheet[position] = column_data

        # Analysis Worksheet
        analysis_worksheet["A1"] = "Price Range"
        analysis_worksheet["B1"] = "Quantity"
        analysis_worksheet["C1"] = "Range Value"
        analysis_worksheet["D1"] = "Estimated Revenue"
        row_number = 2
        for title_range, data in self.price_analysis.items():
            analysis_worksheet["A{}".format(row_number)] = title_range
            analysis_worksheet["B{}".format(row_number)] = data[0]
            analysis_worksheet["C{}".format(row_number)] = data[1]
            estimated_revenue = Constant.working_products_decimal \
                                * Constant.estimated_quantity_demand[title_range] \
                                * data[1]
            analysis_worksheet["D{}".format(row_number)] = estimated_revenue
            row_number += 1
        analysis_worksheet["C9"] = "Lot Price:"
        analysis_worksheet["D9"] = self.lot_price
        analysis_worksheet["C10"] = "Shipping Costs:"
        analysis_worksheet["D10"] = 100
        analysis_worksheet["C11"] = "Total Costs:"
        analysis_worksheet["D11"] = self.lot_price + 100
        analysis_worksheet["C12"] = "Estimated Revenue:"
        analysis_worksheet["D12"] = "=SUM(D2:D8)"
        analysis_worksheet["C13"] = "Estimated Profit:"
        analysis_worksheet["D13"] = "=D12-D11"
        directory = os.path.split(xlsx_location)[0]
        if (directory and not (os.path.exists(directory))):
            os.makedirs(directory)
        workbook.save(filename=xlsx_location)


class LoadingBar:
    bar_length = 80
    total_products = 0
    current_product = 0

    def __init__(self, total_products: int) -> None:
        self.total_products = total_products

    def increment_products(self, increment_by: int = 1):
        self.current_product += increment_by

    def output_loading_bar(self):
        completed_decimal = self.current_product / self.total_products
        completed_percentage = round(completed_decimal * 100)
        current_bar_progress = round(completed_decimal * self.bar_length)
        print("[", end="")
        print("=" * current_bar_progress, end="")
        print(" " * (self.bar_length - current_bar_progress), end="")
        print("]", end="")
        print("\t{}% Completed".format(completed_percentage))


class CommandLineArguments:
    arguments = {"-site": 1,
                 "--s": 1,
                 "-output": 1,
                 "--o": 1,
                 "-list": 1,
                 "--l": 1,
                 "-list_output": 1,
                 "--lo": 1}
    parsed_arguments = {"site": [],
                        "site_output": [],
                        "list": [],
                        "list_output": []}

    def __init__(self, arguments: list):
        arguments = arguments[1:]
        i = 0
        while i < len(arguments):
            if (arguments[i] in self.arguments):
                if (arguments[i].lower() in ["-site", "--s"]):
                    self.parsed_arguments["site"].append(arguments[i + 1])
                elif (arguments[i].lower() in ["-output", "--o"]):
                    self.parsed_arguments["site_output"].append(arguments[i + 1])
                elif (arguments[i].lower in ["-list", "--l"]):
                    self.parsed_arguments["list"].append(arguments[i + 1])
                elif (arguments[i].lower() in ["-list_output", "--lo"]):
                    self.parsed_arguments["list_output"].append(arguments[i + 1])
                i += self.arguments[arguments[i].lower()] + 1
            else:
                print("Unrecognised argument '{}'".format(arguments[i]))
                i += 1


class BritDeals(WebsiteAnalysis):
    def analyse(self,
                data_location: str = None,
                website_currency: Currency = None) -> None:
        if (data_location and os.path.exists(data_location) and os.path.isfile(data_location)):
            workbook = load_workbook(filename=data_location)
            sheet_names = workbook.sheetnames
            products = workbook[sheet_names[0]]

            product_headings = []

            for row in products.rows:
                if (len(product_headings) == 0):
                    for cell in row:
                        product_headings.append(cell.value)
                else:
                    current_product = ProductData(row[0].value,  # ASIN
                                                  row[1].value,  # ASIN Location
                                                  row[2].value,  # Quantity
                                                  row[3].value,  # Condition
                                                  row[4].value,  # Description
                                                  row[5].value,  # Expected Amazon price
                                                  Currency("Great British Pounds",  # Price Currency
                                                           "GBP",
                                                           "£"),
                                                  row[7].value)
                    self.product_data.append(current_product)
        else:
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.implicitly_wait(0)
            driver.get(self.url)

            driver.execute_script("window.open()")
            driver.switch_to.window(driver.window_handles[1])
            driver.get("https://google.com")
            driver.switch_to.window(driver.window_handles[0])

            self.lot_price = float(driver.find_element_by_css_selector("span.f-s-40").text.replace(",", "").lstrip("£"))
            manifest_table = driver.find_element_by_css_selector("table.table.withdraw__table.manifest__table")
            # noinspection SpellCheckingInspection
            table_rows = manifest_table.find_elements_by_css_selector("tbody > tr")

            loading_bar = LoadingBar(len(table_rows))

            for row in table_rows:
                loading_bar.output_loading_bar()
                row_data = row.find_elements_by_css_selector("*")
                price_url = "https://www.amazon.co.uk/dp/{}"
                # noinspection PyPep8Naming
                ASIN = row_data[3].text.split("\n")[0]
                driver.switch_to.window(driver.window_handles[1])
                driver.get(price_url.format(ASIN))
                price_data_list = driver.find_elements_by_css_selector("span#priceblock_ourprice")
                current_amazon_price = 0
                if (len(price_data_list)):
                    current_amazon_price = float(price_data_list[0].text.replace(",", "").lstrip("£"))

                driver.switch_to.window(driver.window_handles[0])
                self.product_data.append(ProductData(ASIN,
                                                     "UK",
                                                     row_data[8].text.split()[0],
                                                     row_data[9].text,
                                                     row_data[13].text.split("\n")[0],
                                                     float(row_data[19].text.replace(",", "").lstrip("£")),
                                                     Currency("Great British Pounds", "GBP", "£"),
                                                     current_amazon_price))
                loading_bar.increment_products()
            loading_bar.output_loading_bar()
        self.price_analysis = Analysis.analyze_amazon_prices(self.product_data)


class GemWholesale(WebsiteAnalysis):
    def analyse(self,
                data_location: str = None,
                website_currency: Currency = None) -> None:
        if (data_location and os.path.exists(data_location) and os.path.isfile(data_location)):
            workbook = load_workbook(filename=data_location)
            sheet_names = workbook.sheetnames
            products = workbook[sheet_names[0]]

            product_headings = []

            for row in products.rows:
                if (len(product_headings) == 0):
                    for cell in row:
                        product_headings.append(cell.value)
                else:
                    current_product = ProductData(row[0].value,  # ASIN
                                                  row[1].value,  # ASIN Location
                                                  row[2].value,  # Quantity
                                                  row[3].value,  # Condition
                                                  row[4].value,  # Description
                                                  row[5].value,  # Expected Amazon price
                                                  Currency("Great British Pounds",  # Price Currency
                                                           "GBP",
                                                           "£"),
                                                  row[7].value)
                    self.product_data.append(current_product)
        else:
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.implicitly_wait(0)
            driver.get(self.url)

            driver.execute_script("window.open()")
            driver.switch_to.window(driver.window_handles[1])
            driver.get("https://google.com")
            driver.switch_to.window(driver.window_handles[0])

            self.lot_price = 0.0
            manifest_table = driver.find_element_by_css_selector("table.xl256612635")
            # noinspection SpellCheckingInspection
            table_rows = manifest_table.find_elements_by_css_selector("tbody > tr")

            loading_bar = LoadingBar(len(table_rows))

            current_row = 0
            for row in table_rows:
                if (current_row > 4):
                    loading_bar.output_loading_bar()
                    row_data = row.find_elements_by_css_selector("*")
                    price_url = "https://www.amazon.co.uk/dp/{}"
                    # noinspection PyPep8Naming
                    ASIN = row_data[0].text
                    driver.switch_to.window(driver.window_handles[1])
                    driver.get(price_url.format(ASIN))
                    price_data_list = driver.find_elements_by_css_selector("span#priceblock_ourprice")
                    current_amazon_price = 0
                    if (len(price_data_list)):
                        current_amazon_price = float(price_data_list[0].text.replace(",", "").lstrip("£"))

                    driver.switch_to.window(driver.window_handles[0])
                    try:
                        self.product_data.append(ProductData(ASIN,
                                                             "UK",
                                                             row_data[3].text,
                                                             "Returns",
                                                             row_data[1].text,
                                                             float(row_data[3].text.replace(",", "").lstrip("£")),
                                                             Currency("Great British Pounds", "GBP", "£"),
                                                             current_amazon_price))
                    except Exception as e:
                        print(e)
                    loading_bar.increment_products()
                else:
                    current_row += 1
            loading_bar.output_loading_bar()
        self.price_analysis = Analysis.analyze_amazon_prices(self.product_data)


if (__name__ == "__main__"):
    parsed_arguments = CommandLineArguments(sys.argv).parsed_arguments
    if (len(parsed_arguments["site"]) > 0):
        for index, site in enumerate(parsed_arguments["site"]):
            if ("britdeals.co.uk" in site):
                data = BritDeals(site)
                data.analyse()
                data.to_xlsx(parsed_arguments["site_output"][index])
            elif ("gemwholesale.co.uk" in site):
                data = GemWholesale(site)
                data.analyse()
                data.to_xlsx(parsed_arguments["site_output"][index])
